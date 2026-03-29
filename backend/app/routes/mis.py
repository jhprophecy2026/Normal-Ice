"""
MIS Report API
GET /api/mis/report?period=weekly|monthly|yearly

Generates a multi-sheet Excel MIS report covering the requested period.

Format reference:
  - IRDAI circular IRDAI/HLT/REG/CIR/101/07/2021 — standardised TPA MIS reporting format
  - Common Indian TPA software export formats (Medi Assist, Health India TPA, Paramount Health Services TPA)
  - NHA (National Health Authority) PMJAY claim data reporting template

Sheets generated:
  1. Case Summary      — one row per case, full financial lifecycle
  2. Pre-Auth Details  — pre-auth form fields + estimated costs
  3. Enhancement Details — all enhancement requests with variance
  4. Discharge & Settlement — final bill breakdown + TPA settlement
"""
import io
import logging
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.supabase_client import get_supabase

logger = logging.getLogger(__name__)
router = APIRouter()

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark navy
_SUBHEAD_FILL = PatternFill("solid", fgColor="2B6CB0")   # blue
_ALT_FILL     = PatternFill("solid", fgColor="EBF4FF")   # light blue stripe
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT   = Font(bold=True, color="1E3A5F", size=12)
_THIN         = Side(style="thin", color="CBD5E0")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",  vertical="center", wrap_text=True)

def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

def _style_data_row(ws, row: int, col_count: int, alt: bool):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = _ALT_FILL
        cell.border  = _BORDER
        cell.alignment = _LEFT

def _add_title(ws, title: str, subtitle: str, col_span: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c = ws.cell(1, 1, title)
    c.font = _TITLE_FONT
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    c2 = ws.cell(2, 1, subtitle)
    c2.font = Font(italic=True, color="718096", size=9)
    c2.alignment = _CENTER
    ws.row_dimensions[2].height = 16

def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _inr(val) -> str:
    if val is None:
        return "—"
    try:
        return f"₹{float(val):,.0f}"
    except Exception:
        return str(val)

def _date(val) -> str:
    if not val:
        return "—"
    try:
        return str(val)[:10]
    except Exception:
        return str(val)

# ---------------------------------------------------------------------------
# Period helper
# ---------------------------------------------------------------------------

def _cutoff(period: str) -> datetime:
    now = datetime.now(timezone.utc)
    if period == "weekly":
        return now - timedelta(days=7)
    elif period == "monthly":
        return now - timedelta(days=30)
    elif period == "yearly":
        return now - timedelta(days=365)
    raise HTTPException(status_code=400, detail="period must be weekly, monthly, or yearly")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_case_summary(wb: Workbook, pre_auths: list, discharges: dict, settlements: dict, enhancements: dict, period_label: str):
    ws = wb.active
    ws.title = "Case Summary"

    headers = [
        "Sr No", "Bill No", "Patient Name", "ABHA ID", "Hospital",
        "Admission Date", "Discharge Date", "Diagnosis", "ICD-10",
        "Pre-Auth Amt (₹)", "Enhancement Amt (₹)", "Final Bill (₹)",
        "Settled Amt (₹)", "Deduction (₹)", "Status", "TAT (Days)"
    ]
    widths = [6, 16, 20, 18, 22, 14, 14, 28, 12, 16, 16, 16, 14, 14, 14, 10]

    _add_title(ws, "MIS Report — Case Summary", period_label, len(headers))
    ws.append([])  # blank row 3
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, pa in enumerate(pre_auths, 1):
        bill_no = pa.get("bill_no", "")
        dis = discharges.get(bill_no, {})
        sett = settlements.get(bill_no, {})
        enhs = enhancements.get(pa["id"], [])

        enh_total = sum(e.get("revised_total_estimated_cost") or 0 for e in enhs)

        # TAT
        tat = "—"
        try:
            adm = pa.get("admission_date") or pa.get("created_at", "")[:10]
            sdate = sett.get("settlement_date") or dis.get("discharge_date") or ""
            if adm and sdate:
                tat = str((datetime.fromisoformat(sdate[:10]) - datetime.fromisoformat(adm[:10])).days)
        except Exception:
            pass

        # Status
        if sett:
            status = "Settled"
        elif dis:
            status = "Discharged"
        elif enhs:
            status = "Enhanced"
        else:
            status = pa.get("status", "Pre-Auth").title()

        row = [
            i, bill_no,
            pa.get("patient_name", "—"),
            pa.get("abha_id", "—"),
            pa.get("hospital_name", "—"),
            _date(pa.get("admission_date")),
            _date(dis.get("discharge_date")),
            pa.get("provisional_diagnosis", "—"),
            pa.get("icd10_diagnosis_code", "—"),
            pa.get("total_estimated_cost"),
            enh_total or None,
            dis.get("total_bill_amount"),
            sett.get("final_settlement_amount"),
            sett.get("deduction_amount"),
            status,
            tat,
        ]
        ws.append(row)
        _style_data_row(ws, 4 + i, len(headers), i % 2 == 0)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A5"


def _sheet_preauth(wb: Workbook, pre_auths: list, period_label: str):
    ws = wb.create_sheet("Pre-Auth Details")

    headers = [
        "Sr No", "Pre-Auth ID", "Bill No", "Patient Name", "ABHA ID",
        "Hospital", "Doctor", "Admission Date", "Admission Type", "Room Type",
        "Diagnosis", "ICD-10", "Treatment Type", "Surgery Name",
        "Exp. Days", "ICU Days", "Estimated Cost (₹)", "Status", "Created Date"
    ]
    widths = [6, 36, 16, 20, 18, 22, 20, 14, 14, 12, 28, 12, 16, 24, 10, 10, 16, 12, 14]

    _add_title(ws, "MIS Report — Pre-Authorization Details", period_label, len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, pa in enumerate(pre_auths, 1):
        row = [
            i,
            pa.get("id", "—"),
            pa.get("bill_no", "—"),
            pa.get("patient_name", "—"),
            pa.get("abha_id", "—"),
            pa.get("hospital_name", "—"),
            pa.get("doctor_name", "—"),
            _date(pa.get("admission_date")),
            pa.get("admission_type", "—"),
            pa.get("room_type", "—"),
            pa.get("provisional_diagnosis", "—"),
            pa.get("icd10_diagnosis_code", "—"),
            pa.get("treatment_medical_management") and "Medical" or pa.get("treatment_surgical") and "Surgical" or "—",
            pa.get("surgery_name", "—"),
            pa.get("expected_days_in_hospital"),
            pa.get("days_in_icu"),
            pa.get("total_estimated_cost"),
            pa.get("status", "—").title(),
            _date(pa.get("created_at")),
        ]
        ws.append(row)
        _style_data_row(ws, 4 + i, len(headers), i % 2 == 0)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A5"


def _sheet_enhancements(wb: Workbook, all_enhs: list, period_label: str):
    ws = wb.create_sheet("Enhancement Details")

    headers = [
        "Sr No", "Enhancement ID", "Bill No", "Pre-Auth ID", "Seq No",
        "Reason", "Updated Diagnosis", "Updated ICD-10",
        "Original Cost (₹)", "Revised Cost (₹)", "Variance (₹)",
        "Status", "Created Date"
    ]
    widths = [6, 36, 16, 36, 8, 30, 28, 12, 16, 16, 14, 12, 14]

    _add_title(ws, "MIS Report — Enhancement Details", period_label, len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, e in enumerate(all_enhs, 1):
        orig = e.get("original_total_cost") or 0
        revised = e.get("revised_total_estimated_cost") or 0
        variance = (revised - orig) if (orig and revised) else None

        row = [
            i,
            e.get("id", "—"),
            e.get("bill_no", "—"),
            e.get("pre_auth_id", "—"),
            e.get("sequence_no"),
            e.get("reason", "—"),
            e.get("updated_diagnosis") or e.get("original_diagnosis", "—"),
            e.get("updated_icd10_code") or e.get("original_icd10_code", "—"),
            orig or None,
            revised or None,
            variance,
            e.get("status", "—").title(),
            _date(e.get("created_at")),
        ]
        ws.append(row)
        _style_data_row(ws, 4 + i, len(headers), i % 2 == 0)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A5"


def _sheet_discharge_settlement(wb: Workbook, pre_auths: list, discharges: dict, settlements: dict, period_label: str):
    ws = wb.create_sheet("Discharge & Settlement")

    headers = [
        "Sr No", "Bill No", "Patient Name", "Discharge Date",
        "Final Diagnosis", "ICD-10 Codes", "Procedure Codes",
        "Room (₹)", "ICU (₹)", "Surgery (₹)", "Medicines (₹)",
        "Investigations (₹)", "Other (₹)", "Total Bill (₹)",
        "Pre-Auth Amt (₹)", "Settled Amt (₹)", "Deduction (₹)",
        "Deduction Reason", "TPA Remarks", "Settlement Date"
    ]
    widths = [6, 16, 20, 14, 28, 14, 14, 12, 12, 12, 12, 14, 10, 14, 14, 14, 12, 24, 24, 14]

    _add_title(ws, "MIS Report — Discharge & Settlement", period_label, len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, pa in enumerate(pre_auths, 1):
        bill_no = pa.get("bill_no", "")
        dis  = discharges.get(bill_no, {})
        sett = settlements.get(bill_no, {})

        row = [
            i,
            bill_no,
            pa.get("patient_name", "—"),
            _date(dis.get("discharge_date")),
            dis.get("final_diagnosis") or pa.get("provisional_diagnosis", "—"),
            dis.get("final_icd10_codes") or pa.get("icd10_diagnosis_code", "—"),
            dis.get("procedure_codes", "—"),
            dis.get("room_charges"),
            dis.get("icu_charges"),
            dis.get("surgery_charges"),
            dis.get("medicine_charges"),
            dis.get("investigation_charges"),
            dis.get("other_charges"),
            dis.get("total_bill_amount"),
            pa.get("total_estimated_cost"),
            sett.get("final_settlement_amount"),
            sett.get("deduction_amount"),
            sett.get("deduction_reason", "—"),
            sett.get("tpa_remarks", "—"),
            _date(sett.get("settlement_date")),
        ]
        ws.append(row)
        _style_data_row(ws, 4 + i, len(headers), i % 2 == 0)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Main endpoint
# ---------------------------------------------------------------------------

@router.get("/mis/report")
async def download_mis_report(period: str = Query("monthly", pattern="^(weekly|monthly|yearly)$")):
    """
    Generate and stream a multi-sheet Excel MIS report for the given period.
    """
    sb = get_supabase()
    cutoff = _cutoff(period)
    cutoff_str = cutoff.isoformat()

    period_labels = {
        "weekly":  "Weekly Report — Last 7 Days",
        "monthly": "Monthly Report — Last 30 Days",
        "yearly":  "Yearly Report — Last 365 Days",
    }
    period_label = f"{period_labels[period]}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"

    # Fetch pre-auths created within period
    pa_res = sb.table("pre_auth_requests") \
        .select("*") \
        .not_.is_("bill_no", "null") \
        .gte("created_at", cutoff_str) \
        .order("created_at", desc=True) \
        .execute()

    pre_auths = pa_res.data or []
    if not pre_auths:
        raise HTTPException(status_code=404, detail=f"No cases found in the selected {period} period.")

    bill_nos     = [r["bill_no"]  for r in pre_auths if r.get("bill_no")]
    pre_auth_ids = [r["id"]       for r in pre_auths]

    # Fetch related records
    dis_res  = sb.table("discharge_requests").select("*").in_("bill_no", bill_nos).execute()     if bill_nos     else None
    sett_res = sb.table("settlement_requests").select("*").in_("bill_no", bill_nos).execute()    if bill_nos     else None
    enh_res  = sb.table("enhancement_requests").select("*").in_("pre_auth_id", pre_auth_ids) \
                 .order("sequence_no").execute()                                                  if pre_auth_ids else None

    discharges  = {r["bill_no"]:     r for r in (dis_res.data  or []) if r.get("bill_no")}
    settlements = {r["bill_no"]:     r for r in (sett_res.data or []) if r.get("bill_no")}

    enhancements_by_pa: dict = {}
    for e in (enh_res.data or []):
        enhancements_by_pa.setdefault(e["pre_auth_id"], []).append(e)

    all_enhs = enh_res.data or []

    # Build workbook
    wb = Workbook()
    _sheet_case_summary(wb, pre_auths, discharges, settlements, enhancements_by_pa, period_label)
    _sheet_preauth(wb, pre_auths, period_label)
    _sheet_enhancements(wb, all_enhs, period_label)
    _sheet_discharge_settlement(wb, pre_auths, discharges, settlements, period_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"MIS_Report_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
